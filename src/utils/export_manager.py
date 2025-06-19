"""Export Manager

Handles exporting search results to various formats including CSV, Excel, and PDF.
Provides data formatting and export functionality.
"""

import csv
import json
import os
import logging
from datetime import datetime
from typing import List, Dict, Optional
from pathlib import Path

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

class ExportManager:
    """Manages data export to various formats"""
    
    def __init__(self):
        self.supported_formats = ['csv', 'json']
        
        if PANDAS_AVAILABLE and OPENPYXL_AVAILABLE:
            self.supported_formats.append('excel')
        
        if REPORTLAB_AVAILABLE:
            self.supported_formats.append('pdf')
    
    def export_data(self, data: List[Dict], file_path: str, format_type: str, 
                   search_params: Optional[Dict] = None) -> bool:
        """Export data to specified format
        
        Args:
            data: List of business data dictionaries
            file_path: Output file path
            format_type: Export format ('csv', 'excel', 'pdf', 'json')
            search_params: Optional search parameters for metadata
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            if format_type.lower() == 'csv':
                return self._export_csv(data, file_path, search_params)
            elif format_type.lower() == 'excel':
                return self._export_excel(data, file_path, search_params)
            elif format_type.lower() == 'pdf':
                return self._export_pdf(data, file_path, search_params)
            elif format_type.lower() == 'json':
                return self._export_json(data, file_path, search_params)
            else:
                raise ValueError(f"Unsupported format: {format_type}")
        except Exception as e:
            print(f"Export error: {e}")
            return False
    
    def _export_csv(self, data: List[Dict], file_path: str, search_params: Optional[Dict] = None) -> bool:
        """Export data to CSV format with improved error handling"""
        if not data:
            return False
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Create backup if file exists
        if os.path.exists(file_path):
            backup_name = f"{file_path}.backup"
            os.rename(file_path, backup_name)
        
        # Get all unique keys from all records
        all_keys = set()
        for record in data:
            all_keys.update(record.keys())
        
        fieldnames = sorted(list(all_keys))
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
            
            # Write header comment with search parameters
            if search_params:
                csvfile.write(f"# Export generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                csvfile.write(f"# Search location: {search_params.get('location', 'N/A')}\n")
                csvfile.write(f"# Business type: {search_params.get('business_type', 'N/A')}\n")
                csvfile.write(f"# Total results: {len(data)}\n")
                csvfile.write("#\n")
            
            writer.writeheader()
            
            # Write data with error handling for individual rows
            for i, record in enumerate(data):
                try:
                    # Convert complex data types to strings
                    row = {}
                    for key, value in record.items():
                        if isinstance(value, (list, dict)):
                            row[key] = json.dumps(value)
                        else:
                            row[key] = value
                    writer.writerow(row)
                except Exception as e:
                    print(f"Error writing row {i}: {e}")
                    continue
        
        return True
    
    def _export_excel(self, data: List[Dict], file_path: str, search_params: Optional[Dict] = None) -> bool:
        """Export data to Excel format"""
        if not PANDAS_AVAILABLE or not OPENPYXL_AVAILABLE:
            raise ImportError("pandas and openpyxl are required for Excel export")
        
        if not data:
            return False
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Business Leads"
        
        # Add metadata sheet if search params provided
        if search_params:
            # Create metadata sheet
            meta_ws = wb.create_sheet("Search Info")
            meta_ws['A1'] = "Export Information"
            meta_ws['A1'].font = Font(bold=True, size=14)
            
            meta_ws['A3'] = "Export Date:"
            meta_ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            meta_ws['A4'] = "Search Location:"
            meta_ws['B4'] = search_params.get('location', 'N/A')
            
            meta_ws['A5'] = "Business Type:"
            meta_ws['B5'] = search_params.get('business_type', 'N/A')
            
            meta_ws['A6'] = "Total Results:"
            meta_ws['B6'] = len(data)
            
            # Style metadata
            for row in meta_ws['A3:A6']:
                for cell in row:
                    cell.font = Font(bold=True)
        
        # Add data to main sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(file_path)
        return True
    
    def _export_pdf(self, data: List[Dict], file_path: str, search_params: Optional[Dict] = None) -> bool:
        """Export data to PDF format"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export")
        
        if not data:
            return False
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Create PDF document
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        story = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add title
        story.append(Paragraph("Business Leads Export", title_style))
        story.append(Spacer(1, 12))
        
        # Add search information if provided
        if search_params:
            info_data = [
                ['Export Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Search Location:', search_params.get('location', 'N/A')],
                ['Business Type:', search_params.get('business_type', 'N/A')],
                ['Total Results:', str(len(data))]
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 3*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.grey),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(info_table)
            story.append(Spacer(1, 20))
        
        # Prepare data for table
        if data:
            # Get all unique keys
            all_keys = set()
            for record in data:
                all_keys.update(record.keys())
            
            headers = sorted(list(all_keys))
            
            # Create table data
            table_data = [headers]
            for record in data:
                row = [str(record.get(key, '')) for key in headers]
                table_data.append(row)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        # Build PDF
        doc.build(story)
        return True
    
    def _export_json(self, data: List[Dict], file_path: str, search_params: Optional[Dict] = None) -> bool:
        """Export data to JSON format with improved error handling"""
        if not data:
            return False
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Create backup if file exists
        if os.path.exists(file_path):
            backup_name = f"{file_path}.backup"
            os.rename(file_path, backup_name)
        
        # Prepare export data with enhanced metadata
        export_data = {
            'export_info': {
                'export_date': datetime.now().isoformat(),
                'total_results': len(data),
                'export_version': '2.0'
            },
            'search_parameters': search_params or {},
            'results': data
        }
        
        # Write JSON file with error handling
        try:
            with open(file_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(export_data, jsonfile, indent=2, ensure_ascii=False, default=str)
        except PermissionError as e:
            print(f"Permission denied writing to {file_path}: {e}")
            return False
        except Exception as e:
            print(f"Error writing JSON file: {e}")
            return False
        
        return True
    
    def _get_standardized_fieldnames(self, businesses):
        """Get standardized fieldnames for CSV export"""
        # Define preferred field order
        preferred_fields = [
            'name', 'website', 'phone', 'email', 'address',
            'performance_score', 'seo_score', 'accessibility_score', 'best_practices_score',
            'priority', 'business_type', 'description'
        ]
        
        # Get all unique fields
        all_fields = set()
        for business in businesses:
            all_fields.update(business.keys())
        
        # Order fields: preferred first, then alphabetical
        ordered_fields = []
        for field in preferred_fields:
            if field in all_fields:
                ordered_fields.append(field)
                all_fields.remove(field)
        
        # Add remaining fields alphabetically
        ordered_fields.extend(sorted(all_fields))
        return ordered_fields
    
    def _prepare_csv_row(self, business, fieldnames):
        """Prepare a business row for CSV export"""
        row = {}
        for field in fieldnames:
            value = business.get(field, '')
            if isinstance(value, (list, dict)):
                row[field] = json.dumps(value, ensure_ascii=False)
            elif value is None:
                row[field] = ''
            else:
                row[field] = str(value)
        return row
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported export formats
        
        Returns:
            List of supported format strings
        """
        return self.supported_formats.copy()
    
    def is_format_supported(self, format_type: str) -> bool:
        """Check if a format is supported
        
        Args:
            format_type: Format to check
            
        Returns:
            True if format is supported
        """
        return format_type.lower() in self.supported_formats
    
    def get_format_extension(self, format_type: str) -> str:
        """Get file extension for format
        
        Args:
            format_type: Format type
            
        Returns:
            File extension including dot
        """
        extensions = {
            'csv': '.csv',
            'excel': '.xlsx',
            'pdf': '.pdf',
            'json': '.json'
        }
        return extensions.get(format_type.lower(), '.txt')
    
    def suggest_filename(self, search_params: Optional[Dict] = None, format_type: str = 'csv') -> str:
        """Suggest a filename based on search parameters
        
        Args:
            search_params: Search parameters
            format_type: Export format
            
        Returns:
            Suggested filename
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if search_params:
            location = search_params.get('location', 'unknown')
            business_type = search_params.get('business_type', 'all')
            
            # Clean up strings for filename
            location = ''.join(c for c in location if c.isalnum() or c in (' ', '-', '_')).strip()
            business_type = ''.join(c for c in business_type if c.isalnum() or c in (' ', '-', '_')).strip()
            
            # Replace spaces with underscores
            location = location.replace(' ', '_')
            business_type = business_type.replace(' ', '_')
            
            filename = f"leads_{location}_{business_type}_{timestamp}"
        else:
            filename = f"business_leads_{timestamp}"
        
        return filename + self.get_format_extension(format_type)